from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi import UploadFile, File
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, extract, and_
from typing import List, Optional, Dict, Any
from datetime import date, datetime
from collections import defaultdict
from fastapi.responses import StreamingResponse
import pandas as pd
from io import BytesIO
from datetime import datetime
from fastapi import UploadFile, File

from ..database import get_db
from ..models.usuario import Usuario
from ..models.paciente import Paciente
from ..models.examen_base import ExamenBase
from ..models.examen_tac import ExamenTAC
from ..models.examen_rx import ExamenRX
from ..models.examen_eco import ExamenECO
from ..models.catalogos import PersonalMedico
from ..middleware.auth_middleware import get_current_user, require_admin, require_ingresador_o_admin
from ..utils.helpers import limpiar_rut
from ..models.catalogos import Prevision, Procedencia, CodigoMAI, Diagnostico, PersonalMedico
from ..models.examen_especifico import ExamenEspecifico

router = APIRouter()

# ============================================
# ESTADÍSTICAS GENERALES
# ============================================

@router.get("/estadisticas-generales")
def estadisticas_generales(
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Estadísticas generales del sistema
    
    Retorna:
    - Total de exámenes por tipo
    - Total de pacientes únicos
    - Promedio de exámenes por día
    - Distribución por tipo de atención
    """
    
    # Query base
    query = db.query(ExamenBase).filter(ExamenBase.deleted_at.is_(None))
    
    # Aplicar filtros de fecha
    if fecha_inicio:
        query = query.filter(ExamenBase.fecha_realizacion >= fecha_inicio)
    if fecha_fin:
        query = query.filter(ExamenBase.fecha_realizacion <= fecha_fin)
    
    # Total de exámenes por tipo
    examenes_por_tipo = db.query(
        ExamenBase.tipo_examen,
        func.count(ExamenBase.id).label('total')
    ).filter(ExamenBase.deleted_at.is_(None))
    
    if fecha_inicio:
        examenes_por_tipo = examenes_por_tipo.filter(ExamenBase.fecha_realizacion >= fecha_inicio)
    if fecha_fin:
        examenes_por_tipo = examenes_por_tipo.filter(ExamenBase.fecha_realizacion <= fecha_fin)
    
    examenes_por_tipo = examenes_por_tipo.group_by(ExamenBase.tipo_examen).all()
    
    # Total de pacientes únicos
    pacientes_unicos = query.with_entities(
        func.count(func.distinct(ExamenBase.paciente_id))
    ).scalar()
    
    # Distribución por atención
    por_atencion = db.query(
        ExamenBase.atencion,
        func.count(ExamenBase.id).label('total')
    ).filter(ExamenBase.deleted_at.is_(None))
    
    if fecha_inicio:
        por_atencion = por_atencion.filter(ExamenBase.fecha_realizacion >= fecha_inicio)
    if fecha_fin:
        por_atencion = por_atencion.filter(ExamenBase.fecha_realizacion <= fecha_fin)
    
    por_atencion = por_atencion.group_by(ExamenBase.atencion).all()
    
    # Construir respuesta
    return {
        "examenes_por_tipo": {
            examen.tipo_examen: examen.total 
            for examen in examenes_por_tipo
        },
        "total_examenes": sum(examen.total for examen in examenes_por_tipo),
        "pacientes_unicos": pacientes_unicos,
        "por_atencion": {
            atencion.atencion: atencion.total 
            for atencion in por_atencion
        }
    }

# ============================================
# EXÁMENES POR PERÍODO (Para gráficos de línea)
# ============================================

@router.get("/por-periodo")
def examenes_por_periodo(
    anio: int,
    tipo_examen: Optional[str] = Query(None, pattern="^(TAC|RX|ECO)$"),
    agrupar_por: str = Query("mes", pattern="^(mes|semana)$"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Obtener cantidad de exámenes agrupados por mes o semana
    
    Útil para gráficos de línea temporal
    """
    
    query = db.query(
        ExamenBase.mes_realizacion,
        func.count(ExamenBase.id).label('total')
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    )
    
    if tipo_examen:
        query = query.filter(ExamenBase.tipo_examen == tipo_examen)
    
    # Agrupar por mes
    resultados = query.group_by(ExamenBase.mes_realizacion).order_by(ExamenBase.mes_realizacion).all()
    
    # Crear array con todos los meses (llenar con 0 los que no tienen datos)
    meses = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
    datos = {mes: 0 for mes in range(1, 13)}
    
    for resultado in resultados:
        datos[resultado.mes_realizacion] = resultado.total
    
    return {
        "labels": meses,
        "data": [datos[mes] for mes in range(1, 13)],
        "anio": anio,
        "tipo_examen": tipo_examen or "Todos"
    }

# ============================================
# COMPARATIVA POR TIPO DE EXAMEN
# ============================================

@router.get("/comparativa-tipos")
def comparativa_tipos(
    anio: int,
    mes: Optional[int] = Query(None, ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Comparar cantidad de exámenes TAC vs RX vs ECO
    
    Útil para gráfico de barras comparativo
    """
    
    query = db.query(
        ExamenBase.tipo_examen,
        func.count(ExamenBase.id).label('total')
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    )
    
    if mes:
        query = query.filter(ExamenBase.mes_realizacion == mes)
    
    resultados = query.group_by(ExamenBase.tipo_examen).all()
    
    datos = {"TAC": 0, "RX": 0, "ECO": 0}
    for resultado in resultados:
        datos[resultado.tipo_examen] = resultado.total
    
    return {
        "labels": ["TAC", "RX", "ECO"],
        "data": [datos["TAC"], datos["RX"], datos["ECO"]],
        "periodo": f"{mes}/{anio}" if mes else str(anio)
    }

# ============================================
# EXÁMENES POR PACIENTE
# ============================================

@router.get("/por-paciente/{paciente_rut}")
def examenes_por_paciente(
    paciente_rut: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Historial completo de exámenes de un paciente
    """
    
    rut_limpio = limpiar_rut(paciente_rut)
    
    # Buscar paciente
    paciente = db.query(Paciente).filter(Paciente.rut == rut_limpio).first()
    
    if not paciente:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Paciente no encontrado"
        )
    
    # Obtener todos sus exámenes
    examenes = db.query(ExamenBase).filter(
        ExamenBase.paciente_id == paciente.id,
        ExamenBase.deleted_at.is_(None)
    ).order_by(ExamenBase.fecha_realizacion.desc()).all()
    
    # Agrupar por tipo
    por_tipo = {"TAC": 0, "RX": 0, "ECO": 0}
    historial = []
    
    for examen in examenes:
        por_tipo[examen.tipo_examen] += 1
        historial.append({
            "id": examen.id,
            "tipo": examen.tipo_examen,
            "fecha": examen.fecha_realizacion.isoformat(),
            "atencion": examen.atencion
        })
    
    return {
        "paciente": {
            "rut": paciente.rut,
            "nombre": paciente.nombre_completo,
            "fecha_nacimiento": paciente.fecha_nacimiento.isoformat() if paciente.fecha_nacimiento else None
        },
        "total_examenes": len(examenes),
        "por_tipo": por_tipo,
        "historial": historial
    }

# ============================================
# TOP MÉDICOS SOLICITANTES (Solo para TAC)
# ============================================

@router.get("/top-medicos")
def top_medicos_solicitantes(
    anio: int,
    mes: Optional[int] = Query(None, ge=1, le=12),
    limit: int = Query(10, ge=1, le=50),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Top médicos que más exámenes TAC solicitan
    """
    
    query = db.query(
        PersonalMedico.nombre,
        func.count(ExamenTAC.id).label('total')
    ).join(
        ExamenTAC, PersonalMedico.id == ExamenTAC.medico_solicitante_id
    ).join(
        ExamenBase, ExamenTAC.examen_base_id == ExamenBase.id
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    )
    
    if mes:
        query = query.filter(ExamenBase.mes_realizacion == mes)
    
    resultados = query.group_by(PersonalMedico.nombre).order_by(
        func.count(ExamenTAC.id).desc()
    ).limit(limit).all()
    
    return {
        "labels": [r.nombre for r in resultados],
        "data": [r.total for r in resultados],
        "periodo": f"{mes}/{anio}" if mes else str(anio)
    }

# ============================================
# DISTRIBUCIÓN POR PREVISIÓN (Gráfico de torta)
# ============================================

@router.get("/por-prevision")
def examenes_por_prevision(
    anio: int,
    mes: Optional[int] = Query(None, ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Distribución de exámenes por tipo de previsión
    
    Útil para gráfico de torta/pie
    """
    
    from ..models.catalogos import Prevision
    
    query = db.query(
        Prevision.nombre,
        func.count(ExamenBase.id).label('total')
    ).join(
        ExamenBase, Prevision.id == ExamenBase.prevision_id
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    )
    
    if mes:
        query = query.filter(ExamenBase.mes_realizacion == mes)
    
    resultados = query.group_by(Prevision.nombre).order_by(
        func.count(ExamenBase.id).desc()
    ).all()
    
    return {
        "labels": [r.nombre for r in resultados],
        "data": [r.total for r in resultados],
        "periodo": f"{mes}/{anio}" if mes else str(anio)
    }

# ============================================
# RESUMEN MENSUAL COMPLETO
# ============================================

@router.get("/resumen-mensual")
def resumen_mensual(
    anio: int,
    mes: int = Query(..., ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """
    Resumen completo del mes con todas las métricas
    """
    
    # Total por tipo
    por_tipo = db.query(
        ExamenBase.tipo_examen,
        func.count(ExamenBase.id).label('total')
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.mes_realizacion == mes,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.tipo_examen).all()
    
    # Por atención
    por_atencion = db.query(
        ExamenBase.atencion,
        func.count(ExamenBase.id).label('total')
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.mes_realizacion == mes,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.atencion).all()
    
    # Por contrato
    por_contrato = db.query(
        ExamenBase.contrato,
        func.count(ExamenBase.id).label('total')
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.mes_realizacion == mes,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.contrato).all()
    
    return {
        "periodo": f"{mes}/{anio}",
        "total_examenes": sum(t.total for t in por_tipo),
        "por_tipo": {t.tipo_examen: t.total for t in por_tipo},
        "por_atencion": {a.atencion: a.total for a in por_atencion},
        "por_contrato": {c.contrato: c.total for c in por_contrato}
    }

# ============================================
# EXPORTAR RESPALDO A EXCEL
# ============================================

@router.get("/exportar-excel")
def exportar_excel(
    anio: Optional[int] = None,
    mes: Optional[int] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
    current_user: Usuario = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Exporta todos los exámenes a Excel con 3 hojas separadas
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Construir query base
    query_tac = db.query(ExamenBase).join(ExamenTAC).options(
        joinedload(ExamenBase.paciente),
        joinedload(ExamenBase.prevision),
        joinedload(ExamenBase.procedencia),
        joinedload(ExamenBase.codigo_mai),
        joinedload(ExamenBase.examen_especifico)
    ).filter(ExamenBase.tipo_examen == "TAC", ExamenBase.deleted_at.is_(None))
    
    query_rx = db.query(ExamenBase).join(ExamenRX).options(
        joinedload(ExamenBase.paciente),
        joinedload(ExamenBase.prevision),
        joinedload(ExamenBase.procedencia),
        joinedload(ExamenBase.codigo_mai),
        joinedload(ExamenBase.examen_especifico)
    ).filter(ExamenBase.tipo_examen == "RX", ExamenBase.deleted_at.is_(None))
    
    query_eco = db.query(ExamenBase).join(ExamenECO).options(
        joinedload(ExamenBase.paciente),
        joinedload(ExamenBase.prevision),
        joinedload(ExamenBase.procedencia),
        joinedload(ExamenBase.codigo_mai),
        joinedload(ExamenBase.examen_especifico)
    ).filter(ExamenBase.tipo_examen == "ECO", ExamenBase.deleted_at.is_(None))
    
    # Aplicar filtros
    if anio:
        query_tac = query_tac.filter(ExamenBase.anio_realizacion == anio)
        query_rx = query_rx.filter(ExamenBase.anio_realizacion == anio)
        query_eco = query_eco.filter(ExamenBase.anio_realizacion == anio)
    
    if mes:
        query_tac = query_tac.filter(ExamenBase.mes_realizacion == mes)
        query_rx = query_rx.filter(ExamenBase.mes_realizacion == mes)
        query_eco = query_eco.filter(ExamenBase.mes_realizacion == mes)
    
    if fecha_inicio:
        query_tac = query_tac.filter(ExamenBase.fecha_realizacion >= fecha_inicio)
        query_rx = query_rx.filter(ExamenBase.fecha_realizacion >= fecha_inicio)
        query_eco = query_eco.filter(ExamenBase.fecha_realizacion >= fecha_inicio)
    
    if fecha_fin:
        query_tac = query_tac.filter(ExamenBase.fecha_realizacion <= fecha_fin)
        query_rx = query_rx.filter(ExamenBase.fecha_realizacion <= fecha_fin)
        query_eco = query_eco.filter(ExamenBase.fecha_realizacion <= fecha_fin)
    
    examenes_tac = query_tac.all()
    examenes_rx = query_rx.all()
    examenes_eco = query_eco.all()
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # ============= HOJA TAC =============
    ws_tac = wb.create_sheet(title="TAC")
    
    # Encabezados TAC
    headers_tac = [
        "ID", "Fecha Realización", "Fecha Solicitud", "Hora", "Nombre", "RUT",
        "F/Nac", "Edad", "Previsión", "Atención", "Procedencia", "Externo",
        "Código", "Examen", #"Protocolo",
        "Cód.ACV", "GES", "M.Contraste",
        "VFGE", "Premedicado", "Diagnóstico", "Médico Sol.", "TM", "Contrato",
        "TP", "Secretaria", "Observación", "Creado el"
    ]
    
    ws_tac.append(headers_tac)
    
    # Aplicar estilo al encabezado
    for cell in ws_tac[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Datos TAC
    for examen_base in examenes_tac:
        examen_tac = db.query(ExamenTAC).options(
#            joinedload(ExamenTAC.protocolo),
            joinedload(ExamenTAC.diagnostico_clinico),
            joinedload(ExamenTAC.medico_solicitante),
            joinedload(ExamenTAC.tm),
            joinedload(ExamenTAC.tp),
            joinedload(ExamenTAC.secretaria)
        ).filter(ExamenTAC.examen_base_id == examen_base.id).first()
        
        row = [
            examen_base.id,
            examen_base.fecha_realizacion.strftime('%Y-%m-%d') if examen_base.fecha_realizacion else '',
            examen_tac.fecha_solicitud.strftime('%Y-%m-%d') if examen_tac.fecha_solicitud else '',
            str(examen_tac.hora_realizacion) if examen_tac.hora_realizacion else '',
            examen_base.paciente.nombre_completo if examen_base.paciente else '',
            examen_base.paciente.rut if examen_base.paciente else '',
            examen_base.paciente.fecha_nacimiento.strftime('%Y-%m-%d') if examen_base.paciente and examen_base.paciente.fecha_nacimiento else '',
            examen_tac.edad if examen_tac.edad else '',
            examen_base.prevision.nombre if examen_base.prevision else '',
            examen_base.atencion,
            examen_base.procedencia.nombre if examen_base.procedencia else '',
            examen_tac.externo if examen_tac.externo else '',
            examen_base.codigo_mai.codigo if examen_base.codigo_mai else '',
            examen_base.examen_especifico.nombre if examen_base.examen_especifico else '',
#            examen_tac.protocolo.nombre if examen_tac.protocolo else '',
            'Sí' if examen_tac.cod_acv else 'No',
            'Sí' if examen_tac.ges else 'No',
            'Sí' if examen_tac.medio_contraste else 'No',
            examen_tac.vfge if examen_tac.vfge else '',
            'Sí' if examen_tac.premedicado else ('No' if examen_tac.premedicado is False else ''),
            examen_tac.diagnostico_clinico.nombre if examen_tac.diagnostico_clinico else '',
            examen_tac.medico_solicitante.nombre if examen_tac.medico_solicitante else '',
            examen_tac.tm.nombre if examen_tac.tm else '',
            examen_base.contrato if examen_base.contrato else '',
            examen_tac.tp.nombre if examen_tac.tp else '',
            examen_tac.secretaria.nombre if examen_tac.secretaria else '',
            examen_tac.observacion if examen_tac.observacion else '',
            examen_base.created_at.strftime('%Y-%m-%d %H:%M') if examen_base.created_at else ''
        ]
        
        ws_tac.append(row)
    
    # ============= HOJA RX =============
    ws_rx = wb.create_sheet(title="RX")
    
    headers_rx = [
        "ID", "Fecha", "Atención", "Previsión", "Procedencia", "RUT", "Nombre",
        "Código", "Examen", "Hora", "Realizado por", "Contrato", "Creado el"
    ]
    
    ws_rx.append(headers_rx)
    
    for cell in ws_rx[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for examen_base in examenes_rx:
        examen_rx = db.query(ExamenRX).options(
            joinedload(ExamenRX.tm_tp)
        ).filter(ExamenRX.examen_base_id == examen_base.id).first()
        
        row = [
            examen_base.id,
            examen_base.fecha_realizacion.strftime('%Y-%m-%d') if examen_base.fecha_realizacion else '',
            examen_base.atencion,
            examen_base.prevision.nombre if examen_base.prevision else '',
            examen_base.procedencia.nombre if examen_base.procedencia else '',
            examen_base.paciente.rut if examen_base.paciente else '',
            examen_base.paciente.nombre_completo if examen_base.paciente else '',
            examen_base.codigo_mai.codigo if examen_base.codigo_mai else '',
            examen_base.examen_especifico.nombre if examen_base.examen_especifico else '',
            str(examen_rx.hora_realizacion) if examen_rx.hora_realizacion else '',
            examen_rx.tm_tp.nombre if examen_rx.tm_tp else '',
            examen_base.contrato if examen_base.contrato else '',
            examen_base.created_at.strftime('%Y-%m-%d %H:%M') if examen_base.created_at else ''
        ]
        
        ws_rx.append(row)
    
    # ============= HOJA ECO =============
    ws_eco = wb.create_sheet(title="ECO")
    
    headers_eco = [
        "ID", "Fecha", "Mes", "RUT", "Nombre", "Atención", "Previsión",
        "Código", "Examen", "Diagnóstico", "Procedencia", "Realizado",
        "Contrato", "Transcribe", "Creado el"
    ]
    
    ws_eco.append(headers_eco)
    
    for cell in ws_eco[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for examen_base in examenes_eco:
        examen_eco = db.query(ExamenECO).options(
            joinedload(ExamenECO.diagnostico),
            joinedload(ExamenECO.realizado),
            joinedload(ExamenECO.transcribe)
        ).filter(ExamenECO.examen_base_id == examen_base.id).first()
        
        row = [
            examen_base.id,
            examen_base.fecha_realizacion.strftime('%Y-%m-%d') if examen_base.fecha_realizacion else '',
            examen_base.mes_realizacion,
            examen_base.paciente.rut if examen_base.paciente else '',
            examen_base.paciente.nombre_completo if examen_base.paciente else '',
            examen_base.atencion,
            examen_base.prevision.nombre if examen_base.prevision else '',
            examen_base.codigo_mai.codigo if examen_base.codigo_mai else '',
            examen_base.examen_especifico.nombre if examen_base.examen_especifico else '',
            examen_eco.diagnostico.nombre if examen_eco.diagnostico else '',
            examen_base.procedencia.nombre if examen_base.procedencia else '',
            examen_eco.realizado.nombre if examen_eco.realizado else '',
            examen_base.contrato if examen_base.contrato else '',
            examen_eco.transcribe.nombre if examen_eco.transcribe else '',
            examen_base.created_at.strftime('%Y-%m-%d %H:%M') if examen_base.created_at else ''
        ]
        
        ws_eco.append(row)
    
    # Ajustar anchos de columna para todas las hojas
    for ws in [ws_tac, ws_rx, ws_eco]:
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=examenes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )




from datetime import time  # Agregar a la línea de datetime
from ..utils.helpers import limpiar_rut  # Ya debería existir

# ============================================
# IMPORTAR EXCEL CON LÓGICA COMPLETA
# ============================================

@router.post("/importar-excel")
async def importar_excel(
    file: UploadFile = File(...),
    current_user: Usuario = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Importa exámenes desde Excel con las siguientes características:
    
    1. Detección de Duplicados:
       - Mismo tipo, fecha, RUT, código y examen específico = duplicado
    
    2. Autocompletado:
       - ID: Siempre nuevo
       - Creado el: Fecha/hora de importación
       - Campos opcionales vacíos: NULL
       - Campos obligatorios vacíos: ERROR
    
    3. Creación de Catálogos:
       - Auto-crear: Previsión, Procedencia, Diagnóstico, Personal Médico, Examen Específico
       - NO auto-crear: Atención (predefinido), Contrato (predefinido), Códigos MAI
    
    4. Reporte de Errores:
       - Descarga Excel con filas rechazadas + columna ERROR
    """
    import pandas as pd
    from io import BytesIO
    from datetime import datetime as dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    from ..models.catalogos import Prevision, Procedencia, CodigoMAI, ExamenEspecifico, Diagnostico, PersonalMedico
    from ..models.paciente import Paciente
    from ..utils.helpers import limpiar_rut
    
    try:
        # ============================================
        # 1. VALIDAR ARCHIVO
        # ============================================
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400, 
                detail="El archivo debe ser formato Excel (.xlsx o .xls)"
            )
        
        # Leer archivo Excel
        contents = await file.read()
        excel_data = BytesIO(contents)
        
        try:
            xls = pd.ExcelFile(excel_data)
            hojas_requeridas = {'TAC', 'RX', 'ECO'}
            hojas_disponibles = set(xls.sheet_names)
            
            if not hojas_requeridas.issubset(hojas_disponibles):
                faltantes = hojas_requeridas - hojas_disponibles
                raise HTTPException(
                    status_code=400, 
                    detail=f"Faltan hojas requeridas: {', '.join(faltantes)}"
                )
        except Exception as e:
            raise HTTPException(
                status_code=400, 
                detail=f"Error al leer Excel: {str(e)}"
            )
        
        # Leer las 3 hojas
        df_tac = pd.read_excel(excel_data, sheet_name='TAC')
        df_rx = pd.read_excel(excel_data, sheet_name='RX')
        df_eco = pd.read_excel(excel_data, sheet_name='ECO')
        
        # ============================================
        # 2. INICIALIZAR CONTADORES Y ERRORES
        # ============================================
        resultados = {
            "TAC": {
                "procesados": 0, 
                "importados": 0, 
                "duplicados": 0, 
                "errores": 0,
                "filas_error": []  # Lista de filas con error para Excel
            },
            "RX": {
                "procesados": 0, 
                "importados": 0, 
                "duplicados": 0, 
                "errores": 0,
                "filas_error": []
            },
            "ECO": {
                "procesados": 0, 
                "importados": 0, 
                "duplicados": 0, 
                "errores": 0,
                "filas_error": []
            }
        }
        
        # ============================================
        # 3. FUNCIONES AUXILIARES
        # ============================================
        
        def verificar_duplicado(tipo: str, fecha: date, rut: str, codigo: str, examen: str) -> bool:
            """Verifica si existe un examen duplicado"""
            try:
                rut_limpio = limpiar_rut(rut)
                
                existe = db.query(ExamenBase).join(Paciente).join(CodigoMAI).join(ExamenEspecifico).filter(
                    ExamenBase.tipo_examen == tipo,
                    ExamenBase.fecha_realizacion == fecha,
                    Paciente.rut == rut_limpio,
                    CodigoMAI.codigo == codigo,
                    ExamenEspecifico.nombre == examen,
                    ExamenBase.deleted_at.is_(None)
                ).first()
                
                return existe is not None
            except:
                return False
        
        def obtener_o_crear_prevision(nombre: str):
            """Busca o crea previsión"""
            if not nombre or pd.isna(nombre):
                return None
            
            nombre = str(nombre).strip()
            prevision = db.query(Prevision).filter(Prevision.nombre == nombre).first()
            
            if not prevision:
                prevision = Prevision(nombre=nombre, activo=True)
                db.add(prevision)
                db.flush()
            
            return prevision
        
        def obtener_o_crear_procedencia(nombre: str):
            """Busca o crea procedencia"""
            if not nombre or pd.isna(nombre):
                return None
            
            nombre = str(nombre).strip()
            procedencia = db.query(Procedencia).filter(Procedencia.nombre == nombre).first()
            
            if not procedencia:
                procedencia = Procedencia(nombre=nombre, activo=True)
                db.add(procedencia)
                db.flush()
            
            return procedencia
        
        def obtener_codigo_mai(codigo: str, tipo: str):
            """Busca código MAI (NO crea si no existe)"""
            if not codigo or pd.isna(codigo):
                return None
            
            codigo = str(codigo).strip()
            return db.query(CodigoMAI).filter(
                CodigoMAI.codigo == codigo,
                CodigoMAI.tipo_examen == tipo
            ).first()
        
        def obtener_o_crear_examen_especifico(nombre: str, tipo: str):
            """Busca o crea examen específico"""
            if not nombre or pd.isna(nombre):
                return None
            
            nombre = str(nombre).strip()
            examen = db.query(ExamenEspecifico).filter(
                ExamenEspecifico.nombre == nombre,
                ExamenEspecifico.tipo_examen == tipo
            ).first()
            
            if not examen:
                examen = ExamenEspecifico(nombre=nombre, tipo_examen=tipo, activo=True)
                db.add(examen)
                db.flush()
            
            return examen
        
        def obtener_o_crear_diagnostico(nombre: str):
            """Busca o crea diagnóstico"""
            if not nombre or pd.isna(nombre):
                return None
            
            nombre = str(nombre).strip()
            diagnostico = db.query(Diagnostico).filter(Diagnostico.nombre == nombre).first()
            
            if not diagnostico:
                diagnostico = Diagnostico(nombre=nombre, activo=True)
                db.add(diagnostico)
                db.flush()
            
            return diagnostico
        
        def obtener_o_crear_personal(nombre: str, tipo: str = 'GENERAL'):
            """Busca o crea personal médico"""
            if not nombre or pd.isna(nombre):
                return None
            
            nombre = str(nombre).strip()
            personal = db.query(PersonalMedico).filter(PersonalMedico.nombre == nombre).first()
            
            if not personal:
                personal = PersonalMedico(nombre=nombre, tipo=tipo, activo=True)
                db.add(personal)
                db.flush()
            
            return personal
        
        def validar_atencion(valor: str) -> tuple[bool, str]:
            """Valida que atención sea uno de los valores predefinidos"""
            if not valor or pd.isna(valor):
                return False, "Atención es obligatoria"
            
            valor = str(valor).strip()
            valores_validos = ['Abierta', 'Cerrada', 'Urgencia']
            
            if valor not in valores_validos:
                return False, f"Atención inválida. Debe ser: {', '.join(valores_validos)}"
            
            return True, valor
        
        def validar_contrato(valor: str) -> tuple[bool, str]:
            """Valida que contrato sea uno de los valores predefinidos o vacío"""
            if not valor or pd.isna(valor):
                return True, None  # Contrato es opcional
            
            valor = str(valor).strip()
            valores_validos = ['Empresa Externa', 'Institucional']
            
            if valor not in valores_validos:
                return False, f"Contrato inválido. Debe ser: {', '.join(valores_validos)} o vacío"
            
            return True, valor
        
        def parsear_fecha(valor) -> tuple[bool, date, str]:
            """Parsea fecha con manejo de errores"""
            if pd.isna(valor):
                return False, None, "Fecha vacía"
            
            try:
                fecha = pd.to_datetime(valor).date()
                return True, fecha, ""
            except:
                return False, None, f"Fecha inválida: {valor}"
        
        def parsear_hora(valor) -> tuple[bool, time, str]:
            """Parsea hora con manejo de errores"""
            if pd.isna(valor):
                return False, None, "Hora vacía"
            
            try:
                if isinstance(valor, time):
                    return True, valor, ""
                
                hora = pd.to_datetime(str(valor), format='%H:%M:%S').time()
                return True, hora, ""
            except:
                try:
                    hora = pd.to_datetime(str(valor)).time()
                    return True, hora, ""
                except:
                    return False, None, f"Hora inválida: {valor}"
        
        def parsear_boolean(valor, nombre_campo: str) -> tuple[bool, bool, str]:
            """Parsea valores booleanos (Sí/No)"""
            if pd.isna(valor):
                return False, None, f"{nombre_campo} es obligatorio"
            
            valor_str = str(valor).strip().lower()
            
            if valor_str in ['sí', 'si', 'yes', 'true', '1', 's']:
                return True, True, ""
            elif valor_str in ['no', 'false', '0', 'n']:
                return True, False, ""
            else:
                return False, None, f"{nombre_campo} inválido: debe ser Sí o No"
        
        # ============================================
        # 4. PROCESAR TAC
        # ============================================
        for idx, row in df_tac.iterrows():
            resultados["TAC"]["procesados"] += 1
            fila_num = idx + 2  # +2 porque Excel empieza en 1 y tiene header
            error_msg = ""
            
            try:
                # --- Validar campos obligatorios ---
                if pd.isna(row.get('Fecha Realización')):
                    error_msg = "Fecha Realización es obligatoria"
                    raise ValueError(error_msg)
                
                if pd.isna(row.get('RUT')):
                    error_msg = "RUT es obligatorio"
                    raise ValueError(error_msg)
                
                if pd.isna(row.get('Nombre')):
                    error_msg = "Nombre es obligatorio"
                    raise ValueError(error_msg)
                
                # --- Parsear fecha ---
                ok, fecha_realizacion, err = parsear_fecha(row['Fecha Realización'])
                if not ok:
                    error_msg = err
                    raise ValueError(error_msg)
                
                # --- Validar RUT ---
                try:
                    rut_paciente = limpiar_rut(str(row['RUT']))
                except:
                    error_msg = f"RUT inválido: {row['RUT']}"
                    raise ValueError(error_msg)
                
                # --- Validar Código MAI ---
                codigo_str = str(row.get('Código', '')).strip() if pd.notna(row.get('Código')) else ''
                examen_str = str(row.get('Examen', '')).strip() if pd.notna(row.get('Examen')) else ''
                
                if not codigo_str:
                    error_msg = "Código MAI es obligatorio"
                    raise ValueError(error_msg)
                
                if not examen_str:
                    error_msg = "Examen Específico es obligatorio"
                    raise ValueError(error_msg)
                
                # --- Verificar duplicado ---
                if verificar_duplicado("TAC", fecha_realizacion, rut_paciente, codigo_str, examen_str):
                    resultados["TAC"]["duplicados"] += 1
                    continue
                
                # --- Validar Atención ---
                ok, atencion = validar_atencion(row.get('Atención'))
                if not ok:
                    error_msg = atencion  # El mensaje de error está en el segundo valor
                    raise ValueError(error_msg)
                
                # --- Validar Contrato ---
                ok, contrato = validar_contrato(row.get('Contrato'))
                if not ok:
                    error_msg = contrato
                    raise ValueError(error_msg)
                
                # --- Parsear Fecha Solicitud ---
                ok, fecha_solicitud, err = parsear_fecha(row.get('Fecha Solicitud'))
                if not ok:
                    error_msg = f"Fecha Solicitud {err}"
                    raise ValueError(error_msg)
                
                # --- Parsear Hora ---
                ok, hora, err = parsear_hora(row.get('Hora'))
                if not ok:
                    error_msg = f"Hora {err}"
                    raise ValueError(error_msg)
                
                # --- Parsear Booleanos ---
                ok, cod_acv, err = parsear_boolean(row.get('Cód.ACV'), 'Cód.ACV')
                if not ok:
                    error_msg = err
                    raise ValueError(error_msg)
                
                ok, ges, err = parsear_boolean(row.get('GES'), 'GES')
                if not ok:
                    error_msg = err
                    raise ValueError(error_msg)
                
                ok, medio_contraste, err = parsear_boolean(row.get('M.Contraste'), 'M.Contraste')
                if not ok:
                    error_msg = err
                    raise ValueError(error_msg)
                
                # --- Buscar o crear Paciente ---
                paciente = db.query(Paciente).filter(Paciente.rut == rut_paciente).first()
                if not paciente:
                    # Parsear fecha nacimiento si existe
                    fecha_nac = None
                    if pd.notna(row.get('F/Nac')):
                        ok, fecha_nac, _ = parsear_fecha(row['F/Nac'])
                    
                    paciente = Paciente(
                        rut=rut_paciente,
                        nombre_completo=str(row['Nombre']).strip(),
                        fecha_nacimiento=fecha_nac
                    )
                    db.add(paciente)
                    db.flush()
                
                # --- Buscar/crear catálogos ---
                prevision = obtener_o_crear_prevision(row.get('Previsión'))
                procedencia = obtener_o_crear_procedencia(row.get('Procedencia'))
                
                codigo_mai = obtener_codigo_mai(codigo_str, "TAC")
                if not codigo_mai:
                    error_msg = f"Código MAI '{codigo_str}' no existe en el sistema"
                    raise ValueError(error_msg)
                
                examen_especifico = obtener_o_crear_examen_especifico(examen_str, "TAC")
                diagnostico = obtener_o_crear_diagnostico(row.get('Diagnóstico'))
                medico = obtener_o_crear_personal(row.get('Médico Sol.'), 'MEDICO')
                tm = obtener_o_crear_personal(row.get('TM'), 'TM')
                tp = obtener_o_crear_personal(row.get('TP'), 'TP')
                secretaria = obtener_o_crear_personal(row.get('Secretaria'), 'SECRETARIA')
                
                # --- Crear ExamenBase ---
                examen_base = ExamenBase(
                    tipo_examen="TAC",
                    fecha_realizacion=fecha_realizacion,
                    atencion=atencion,
                    prevision_id=prevision.id if prevision else None,
                    procedencia_id=procedencia.id if procedencia else None,
                    paciente_id=paciente.id,
                    codigo_mai_id=codigo_mai.id,
                    examen_especifico_id=examen_especifico.id,
                    contrato=contrato,
                    mes_realizacion=fecha_realizacion.month,
                    anio_realizacion=fecha_realizacion.year,
                    created_by=current_user.id,
                    created_at=dt.utcnow()
                )
                db.add(examen_base)
                db.flush()
                
                # --- Parsear edad ---
                edad = None
                if pd.notna(row.get('Edad')):
                    try:
                        edad = int(row['Edad'])
                    except:
                        pass
                
                # --- Parsear VFGE y Premedicado ---
                vfge = str(row.get('VFGE', '')).strip() if pd.notna(row.get('VFGE')) else None
                premedicado = None
                if pd.notna(row.get('Premedicado')):
                    ok, premedicado, _ = parsear_boolean(row['Premedicado'], 'Premedicado')
                
                # --- Crear ExamenTAC ---
                examen_tac = ExamenTAC(
                    examen_base_id=examen_base.id,
                    fecha_solicitud=fecha_solicitud,
                    hora_realizacion=hora,
                    fecha_nacimiento=paciente.fecha_nacimiento,
                    edad=edad,
                    externo=str(row.get('Externo', '')).strip() if pd.notna(row.get('Externo')) else None,
                    cod_acv=cod_acv,
                    ges=ges,
                    medio_contraste=medio_contraste,
                    vfge=vfge,
                    premedicado=premedicado,
                    diagnostico_clinico_id=diagnostico.id if diagnostico else None,
                    medico_solicitante_id=medico.id if medico else None,
                    tm_id=tm.id if tm else None,
                    tp_id=tp.id if tp else None,
                    secretaria_id=secretaria.id if secretaria else None,
                    observacion=str(row.get('Observación', '')).strip() if pd.notna(row.get('Observación')) else None
                )
                db.add(examen_tac)
                
                resultados["TAC"]["importados"] += 1
                
            except Exception as e:
                resultados["TAC"]["errores"] += 1
                
                # Agregar fila a lista de errores
                fila_error = row.to_dict()
                fila_error['FILA'] = fila_num
                fila_error['ERROR'] = error_msg if error_msg else str(e)
                resultados["TAC"]["filas_error"].append(fila_error)
                
                continue
        
        # ============================================
        # 5. PROCESAR RX
        # ============================================
        for idx, row in df_rx.iterrows():
            resultados["RX"]["procesados"] += 1
            fila_num = idx + 2
            error_msg = ""
            
            try:
                # --- Validar campos obligatorios ---
                if pd.isna(row.get('Fecha')):
                    error_msg = "Fecha es obligatoria"
                    raise ValueError(error_msg)
                
                if pd.isna(row.get('RUT')):
                    error_msg = "RUT es obligatorio"
                    raise ValueError(error_msg)
                
                if pd.isna(row.get('Nombre')):
                    error_msg = "Nombre es obligatorio"
                    raise ValueError(error_msg)
                
                # --- Parsear fecha ---
                ok, fecha_realizacion, err = parsear_fecha(row['Fecha'])
                if not ok:
                    error_msg = err
                    raise ValueError(error_msg)
                
                # --- Validar RUT ---
                try:
                    rut_paciente = limpiar_rut(str(row['RUT']))
                except:
                    error_msg = f"RUT inválido: {row['RUT']}"
                    raise ValueError(error_msg)
                
                # --- Validar campos obligatorios ---
                codigo_str = str(row.get('Código', '')).strip() if pd.notna(row.get('Código')) else ''
                examen_str = str(row.get('Examen', '')).strip() if pd.notna(row.get('Examen')) else ''
                
                if not codigo_str:
                    error_msg = "Código MAI es obligatorio"
                    raise ValueError(error_msg)
                
                if not examen_str:
                    error_msg = "Examen Específico es obligatorio"
                    raise ValueError(error_msg)
                
                # --- Verificar duplicado ---
                if verificar_duplicado("RX", fecha_realizacion, rut_paciente, codigo_str, examen_str):
                    resultados["RX"]["duplicados"] += 1
                    continue
                
                # --- Validar Atención ---
                ok, atencion = validar_atencion(row.get('Atención'))
                if not ok:
                    error_msg = atencion
                    raise ValueError(error_msg)
                
                # --- Validar Contrato ---
                ok, contrato = validar_contrato(row.get('Contrato'))
                if not ok:
                    error_msg = contrato
                    raise ValueError(error_msg)
                
                # --- Parsear Hora ---
                ok, hora, err = parsear_hora(row.get('Hora'))
                if not ok:
                    error_msg = f"Hora {err}"
                    raise ValueError(error_msg)
                
                # --- Buscar o crear Paciente ---
                paciente = db.query(Paciente).filter(Paciente.rut == rut_paciente).first()
                if not paciente:
                    paciente = Paciente(
                        rut=rut_paciente,
                        nombre_completo=str(row['Nombre']).strip()
                    )
                    db.add(paciente)
                    db.flush()
                
                # --- Buscar/crear catálogos ---
                prevision = obtener_o_crear_prevision(row.get('Previsión'))
                procedencia = obtener_o_crear_procedencia(row.get('Procedencia'))
                
                codigo_mai = obtener_codigo_mai(codigo_str, "RX")
                if not codigo_mai:
                    error_msg = f"Código MAI '{codigo_str}' no existe en el sistema"
                    raise ValueError(error_msg)
                
                examen_especifico = obtener_o_crear_examen_especifico(examen_str, "RX")
                tm_tp = obtener_o_crear_personal(row.get('Realizado por'), 'GENERAL')
                
                # --- Crear ExamenBase ---
                examen_base = ExamenBase(
                    tipo_examen="RX",
                    fecha_realizacion=fecha_realizacion,
                    atencion=atencion,
                    prevision_id=prevision.id if prevision else None,
                    procedencia_id=procedencia.id if procedencia else None,
                    paciente_id=paciente.id,
                    codigo_mai_id=codigo_mai.id,
                    examen_especifico_id=examen_especifico.id,
                    contrato=contrato,
                    mes_realizacion=fecha_realizacion.month,
                    anio_realizacion=fecha_realizacion.year,
                    created_by=current_user.id,
                    created_at=dt.utcnow()
                )
                db.add(examen_base)
                db.flush()
                
                # --- Crear ExamenRX ---
                examen_rx = ExamenRX(
                    examen_base_id=examen_base.id,
                    hora_realizacion=hora,
                    tm_tp_id=tm_tp.id if tm_tp else None
                )
                db.add(examen_rx)
                
                resultados["RX"]["importados"] += 1
                
            except Exception as e:
                resultados["RX"]["errores"] += 1
                
                fila_error = row.to_dict()
                fila_error['FILA'] = fila_num
                fila_error['ERROR'] = error_msg if error_msg else str(e)
                resultados["RX"]["filas_error"].append(fila_error)
                
                continue
        
        # ============================================
        # 6. PROCESAR ECO
        # ============================================
        for idx, row in df_eco.iterrows():
            resultados["ECO"]["procesados"] += 1
            fila_num = idx + 2
            error_msg = ""
            
            try:
                # --- Validar campos obligatorios ---
                if pd.isna(row.get('Fecha')):
                    error_msg = "Fecha es obligatoria"
                    raise ValueError(error_msg)
                
                if pd.isna(row.get('RUT')):
                    error_msg = "RUT es obligatorio"
                    raise ValueError(error_msg)
                
                if pd.isna(row.get('Nombre')):
                    error_msg = "Nombre es obligatorio"
                    raise ValueError(error_msg)
                
                # --- Parsear fecha ---
                ok, fecha_realizacion, err = parsear_fecha(row['Fecha'])
                if not ok:
                    error_msg = err
                    raise ValueError(error_msg)
                
                # --- Validar RUT ---
                try:
                    rut_paciente = limpiar_rut(str(row['RUT']))
                except:
                    error_msg = f"RUT inválido: {row['RUT']}"
                    raise ValueError(error_msg)
                
                # --- Validar campos obligatorios ---
                codigo_str = str(row.get('Código', '')).strip() if pd.notna(row.get('Código')) else ''
                examen_str = str(row.get('Examen', '')).strip() if pd.notna(row.get('Examen')) else ''
                
                if not codigo_str:
                    error_msg = "Código MAI es obligatorio"
                    raise ValueError(error_msg)
                
                if not examen_str:
                    error_msg = "Examen Específico es obligatorio"
                    raise ValueError(error_msg)
                
                # --- Verificar duplicado ---
                if verificar_duplicado("ECO", fecha_realizacion, rut_paciente, codigo_str, examen_str):
                    resultados["ECO"]["duplicados"] += 1
                    continue
                
                # --- Validar Atención ---
                ok, atencion = validar_atencion(row.get('Atención'))
                if not ok:
                    error_msg = atencion
                    raise ValueError(error_msg)
                
                # --- Validar Contrato ---
                ok, contrato = validar_contrato(row.get('Contrato'))
                if not ok:
                    error_msg = contrato
                    raise ValueError(error_msg)
                
                # --- Buscar o crear Paciente ---
                paciente = db.query(Paciente).filter(Paciente.rut == rut_paciente).first()
                if not paciente:
                    paciente = Paciente(
                        rut=rut_paciente,
                        nombre_completo=str(row['Nombre']).strip()
                    )
                    db.add(paciente)
                    db.flush()
                
                # --- Buscar/crear catálogos ---
                prevision = obtener_o_crear_prevision(row.get('Previsión'))
                procedencia = obtener_o_crear_procedencia(row.get('Procedencia'))
                
                codigo_mai = obtener_codigo_mai(codigo_str, "ECO")
                if not codigo_mai:
                    error_msg = f"Código MAI '{codigo_str}' no existe en el sistema"
                    raise ValueError(error_msg)
                
                examen_especifico = obtener_o_crear_examen_especifico(examen_str, "ECO")
                diagnostico = obtener_o_crear_diagnostico(row.get('Diagnóstico'))
                realizado = obtener_o_crear_personal(row.get('Realizado'), 'GENERAL')
                transcribe = obtener_o_crear_personal(row.get('Transcribe'), 'GENERAL')
                
                # --- Crear ExamenBase ---
                examen_base = ExamenBase(
                    tipo_examen="ECO",
                    fecha_realizacion=fecha_realizacion,
                    atencion=atencion,
                    prevision_id=prevision.id if prevision else None,
                    procedencia_id=procedencia.id if procedencia else None,
                    paciente_id=paciente.id,
                    codigo_mai_id=codigo_mai.id,
                    examen_especifico_id=examen_especifico.id,
                    contrato=contrato,
                    mes_realizacion=fecha_realizacion.month,
                    anio_realizacion=fecha_realizacion.year,
                    created_by=current_user.id,
                    created_at=dt.utcnow()
                )
                db.add(examen_base)
                db.flush()
                
                # --- Crear ExamenECO ---
                examen_eco = ExamenECO(
                    examen_base_id=examen_base.id,
                    diagnostico_id=diagnostico.id if diagnostico else None,
                    realizado_id=realizado.id if realizado else None,
                    transcribe_id=transcribe.id if transcribe else None
                )
                db.add(examen_eco)
                
                resultados["ECO"]["importados"] += 1
                
            except Exception as e:
                resultados["ECO"]["errores"] += 1
                
                fila_error = row.to_dict()
                fila_error['FILA'] = fila_num
                fila_error['ERROR'] = error_msg if error_msg else str(e)
                resultados["ECO"]["filas_error"].append(fila_error)
                
                continue
        
        # ============================================
        # 7. COMMIT DE TRANSACCIÓN
        # ============================================
        db.commit()
        
        # ============================================
        # 8. GENERAR EXCEL DE ERRORES (SI HAY)
        # ============================================
        excel_errores = None
        total_errores = resultados["TAC"]["errores"] + resultados["RX"]["errores"] + resultados["ECO"]["errores"]
        
        if total_errores > 0:
            wb_errores = Workbook()
            wb_errores.remove(wb_errores.active)
            
            # Estilos
            error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # --- Hoja TAC Errores ---
            if resultados["TAC"]["filas_error"]:
                ws_tac_err = wb_errores.create_sheet(title="TAC - Errores")
                
                # Headers
                if resultados["TAC"]["filas_error"]:
                    headers = ['FILA', 'ERROR'] + list(resultados["TAC"]["filas_error"][0].keys())
                    headers = [h for h in headers if h not in ['FILA', 'ERROR']] # Remover duplicados
                    headers = ['FILA', 'ERROR'] + headers
                    ws_tac_err.append(headers)
                    
                    # Estilo header
                    for cell in ws_tac_err[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    # Datos
                    for fila_err in resultados["TAC"]["filas_error"]:
                        row_data = [fila_err.get('FILA', ''), fila_err.get('ERROR', '')]
                        for key in headers[2:]:  # Skip FILA y ERROR
                            row_data.append(fila_err.get(key, ''))
                        ws_tac_err.append(row_data)
                        
                        # Marcar fila de error en rojo
                        for cell in ws_tac_err[ws_tac_err.max_row]:
                            if cell.column == 2:  # Columna ERROR
                                cell.fill = error_fill
                                cell.font = Font(color="FFFFFF")
            
            # --- Hoja RX Errores ---
            if resultados["RX"]["filas_error"]:
                ws_rx_err = wb_errores.create_sheet(title="RX - Errores")
                
                if resultados["RX"]["filas_error"]:
                    headers = ['FILA', 'ERROR'] + list(resultados["RX"]["filas_error"][0].keys())
                    headers = [h for h in headers if h not in ['FILA', 'ERROR']]
                    headers = ['FILA', 'ERROR'] + headers
                    ws_rx_err.append(headers)
                    
                    for cell in ws_rx_err[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for fila_err in resultados["RX"]["filas_error"]:
                        row_data = [fila_err.get('FILA', ''), fila_err.get('ERROR', '')]
                        for key in headers[2:]:
                            row_data.append(fila_err.get(key, ''))
                        ws_rx_err.append(row_data)
                        
                        for cell in ws_rx_err[ws_rx_err.max_row]:
                            if cell.column == 2:
                                cell.fill = error_fill
                                cell.font = Font(color="FFFFFF")
            
            # --- Hoja ECO Errores ---
            if resultados["ECO"]["filas_error"]:
                ws_eco_err = wb_errores.create_sheet(title="ECO - Errores")
                
                if resultados["ECO"]["filas_error"]:
                    headers = ['FILA', 'ERROR'] + list(resultados["ECO"]["filas_error"][0].keys())
                    headers = [h for h in headers if h not in ['FILA', 'ERROR']]
                    headers = ['FILA', 'ERROR'] + headers
                    ws_eco_err.append(headers)
                    
                    for cell in ws_eco_err[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                    
                    for fila_err in resultados["ECO"]["filas_error"]:
                        row_data = [fila_err.get('FILA', ''), fila_err.get('ERROR', '')]
                        for key in headers[2:]:
                            row_data.append(fila_err.get(key, ''))
                        ws_eco_err.append(row_data)
                        
                        for cell in ws_eco_err[ws_eco_err.max_row]:
                            if cell.column == 2:
                                cell.fill = error_fill
                                cell.font = Font(color="FFFFFF")
            
            # Guardar en memoria
            output_errores = BytesIO()
            wb_errores.save(output_errores)
            output_errores.seek(0)
            excel_errores = output_errores.getvalue()
        
        # ============================================
        # 9. PREPARAR RESPUESTA
        # ============================================
        return {
            "success": True,
            "resultados": {
                "TAC": {
                    "procesados": resultados["TAC"]["procesados"],
                    "importados": resultados["TAC"]["importados"],
                    "duplicados": resultados["TAC"]["duplicados"],
                    "errores": resultados["TAC"]["errores"]
                },
                "RX": {
                    "procesados": resultados["RX"]["procesados"],
                    "importados": resultados["RX"]["importados"],
                    "duplicados": resultados["RX"]["duplicados"],
                    "errores": resultados["RX"]["errores"]
                },
                "ECO": {
                    "procesados": resultados["ECO"]["procesados"],
                    "importados": resultados["ECO"]["importados"],
                    "duplicados": resultados["ECO"]["duplicados"],
                    "errores": resultados["ECO"]["errores"]
                }
            },
            "tiene_errores": total_errores > 0,
            "excel_errores_base64": excel_errores.hex() if excel_errores else None
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al importar: {str(e)}")
