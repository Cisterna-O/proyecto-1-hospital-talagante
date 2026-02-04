from fastapi import APIRouter, Depends, Query, HTTPException, Path
from fastapi.responses import StreamingResponse
import pandas as pd
from io import BytesIO
from sqlalchemy.orm import Session
from sqlalchemy import func, case, and_
from typing import Optional
from datetime import date, datetime

from ..database import get_db
from ..models.usuario import Usuario
from ..models.examen_base import ExamenBase
from ..models.examen_tac import ExamenTAC
from ..models.catalogos import CodigoMAI, Procedencia, Prevision
from ..middleware.auth_middleware import get_current_user, require_admin

router = APIRouter()

# Definir array de meses GLOBALMENTE al inicio
MESES = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
         'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']

@router.get("/codigo-atencion-mes-actual/{tipo_examen}")
def codigo_atencion_mes_actual(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    hoy = date.today()
    mes = hoy.month
    anio = hoy.year

    ultimo = db.query(ExamenBase).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    ).order_by(ExamenBase.fecha_realizacion.desc()).first()

    if ultimo and (ultimo.mes_realizacion != mes or ultimo.anio_realizacion != anio):
        mes = ultimo.mes_realizacion
        anio = ultimo.anio_realizacion

    query = db.query(
        CodigoMAI.codigo,
        ExamenBase.atencion,
        func.count(ExamenBase.id).label("total")
    ).join(CodigoMAI).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.mes_realizacion == mes,
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(CodigoMAI.codigo, ExamenBase.atencion)

    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.codigo not in tabla:
            tabla[r.codigo] = {}
        tabla[r.codigo][r.atencion] = r.total

    return {
        "titulo": f"Código x Atención - {tipo_examen} - {mes}/{anio}",
        "mes": mes,
        "anio": anio,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/procedencia-atencion-mes-actual/{tipo_examen}")
def procedencia_atencion_mes_actual(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    hoy = date.today()
    mes = hoy.month
    anio = hoy.year

    ultimo = db.query(ExamenBase).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    ).order_by(ExamenBase.fecha_realizacion.desc()).first()

    if ultimo and (ultimo.mes_realizacion != mes or ultimo.anio_realizacion != anio):
        mes = ultimo.mes_realizacion
        anio = ultimo.anio_realizacion

    query = db.query(
        Procedencia.nombre,
        ExamenBase.atencion,
        func.count(ExamenBase.id).label("total")
    ).join(Procedencia).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.mes_realizacion == mes,
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(Procedencia.nombre, ExamenBase.atencion)

    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.nombre not in tabla:
            tabla[r.nombre] = {}
        tabla[r.nombre][r.atencion] = r.total

    return {
        "titulo": f"Procedencia x Atención - {tipo_examen} - {mes}/{anio}",
        "mes": mes,
        "anio": anio,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/codigo-prevision-mes-actual/{tipo_examen}")
def codigo_prevision_mes_actual(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    hoy = date.today()
    mes = hoy.month
    anio = hoy.year

    ultimo = db.query(ExamenBase).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    ).order_by(ExamenBase.fecha_realizacion.desc()).first()

    if ultimo and (ultimo.mes_realizacion != mes or ultimo.anio_realizacion != anio):
        mes = ultimo.mes_realizacion
        anio = ultimo.anio_realizacion

    query = db.query(
        CodigoMAI.codigo,
        Prevision.nombre,
        func.count(ExamenBase.id).label("total")
    ).join(CodigoMAI).join(Prevision).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.mes_realizacion == mes,
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(CodigoMAI.codigo, Prevision.nombre)

    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.codigo not in tabla:
            tabla[r.codigo] = {}
        tabla[r.codigo][r.nombre] = r.total

    return {
        "titulo": f"Código x Previsión - {tipo_examen} - {mes}/{anio}",
        "mes": mes,
        "anio": anio,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/atencion-contrato-mes-actual/{tipo_examen}")
def atencion_contrato_mes_actual(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    hoy = date.today()
    mes = hoy.month
    anio = hoy.year

    ultimo = db.query(ExamenBase).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    ).order_by(ExamenBase.fecha_realizacion.desc()).first()

    if ultimo and (ultimo.mes_realizacion != mes or ultimo.anio_realizacion != anio):
        mes = ultimo.mes_realizacion
        anio = ultimo.anio_realizacion

    query = db.query(
        ExamenBase.atencion,
        ExamenBase.contrato,
        func.count(ExamenBase.id).label("total")
    ).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.mes_realizacion == mes,
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.atencion, ExamenBase.contrato)

    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.atencion not in tabla:
            tabla[r.atencion] = {}
        tabla[r.atencion][r.contrato] = r.total

    return {
        "titulo": f"Atención x Contrato - {tipo_examen} - {mes}/{anio}",
        "mes": mes,
        "anio": anio,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/mes-anio/{tipo_examen}")
def mes_anio_tipo_examen(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        ExamenBase.anio_realizacion,
        ExamenBase.mes_realizacion,
        func.count(ExamenBase.id).label("total")
    ).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.anio_realizacion, ExamenBase.mes_realizacion).order_by(
        ExamenBase.anio_realizacion, ExamenBase.mes_realizacion
    )

    resultados = query.all()

    tabla = {}
    for r in resultados:
        # Usar nombre de mes como columna
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if r.anio_realizacion not in tabla:
            tabla[r.anio_realizacion] = {}
        tabla[r.anio_realizacion][mes_nombre] = r.total

    return {
        "titulo": f"Mes x Año - {tipo_examen}",
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/mes-mc-anio-actual")
def mes_anio_actual(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Gráfico obligatorio: Mes x Medio Contraste - Año Actual"""
    anio_actual = datetime.now().year

    # Query: Mes x Si tiene MC / No tiene MC
    query = db.query(
        ExamenBase.mes_realizacion,
        ExamenTAC.medio_contraste,
        func.count(ExamenBase.id).label('total')
    ).join(ExamenTAC).filter(
        ExamenBase.tipo_examen == "TAC",
        ExamenBase.anio_realizacion == anio_actual,
        ExamenBase.deleted_at.is_(None)
    ).group_by(
        ExamenBase.mes_realizacion,
        ExamenTAC.medio_contraste
    ).order_by(ExamenBase.mes_realizacion)
    
    resultados = query.all()
    
    # Estructura: {mes_nombre: {"Con MC": X, "Sin MC": Y}}
    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if mes_nombre not in tabla:
            tabla[mes_nombre] = {}
        mc_label = "Con MC" if r.medio_contraste else "Sin MC"
        tabla[mes_nombre][mc_label] = r.total
    
    return {
        "titulo": f"Mes x Medio Contraste - {anio_actual}",
        "anio": anio_actual,
        "tabla": tabla
    }

@router.get("/mes-anio-mc")
def mes_anio_mc(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Gráfico obligatorio: Mes x Año - Medio Contraste (todos los años)"""
    # Query: Año x Mes (solo exámenes CON medio contraste)
    query = db.query(
        ExamenBase.anio_realizacion,
        ExamenBase.mes_realizacion,
        func.count(ExamenTAC.id).label('total')
    ).join(ExamenTAC).filter(
        ExamenBase.tipo_examen == "TAC",
        ExamenTAC.medio_contraste == True,
        ExamenBase.deleted_at.is_(None)
    ).group_by(
        ExamenBase.anio_realizacion,
        ExamenBase.mes_realizacion
    ).order_by(
        ExamenBase.anio_realizacion,
        ExamenBase.mes_realizacion
    )
    
    resultados = query.all()
    
    # Estructura: {año: {mes_nombre: total}}
    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if r.anio_realizacion not in tabla:
            tabla[r.anio_realizacion] = {}
        tabla[r.anio_realizacion][mes_nombre] = r.total

    return {
        "titulo": "Mes x Año - Medio Contraste",
        "tabla": tabla
    }

@router.get("/mes-tipo-examen-anio-actual")
def mes_tipo_examen_anio_actual(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    anio = date.today().year

    ultimo = db.query(ExamenBase).filter(
        ExamenBase.deleted_at.is_(None)
    ).order_by(ExamenBase.fecha_realizacion.desc()).first()

    if ultimo and ultimo.anio_realizacion != anio:
        anio = ultimo.anio_realizacion

    query = db.query(
        ExamenBase.mes_realizacion,
        ExamenBase.tipo_examen,
        func.count(ExamenBase.id).label("total")
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.mes_realizacion, ExamenBase.tipo_examen).order_by(
        ExamenBase.mes_realizacion
    )

    resultados = query.all()

    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if mes_nombre not in tabla:
            tabla[mes_nombre] = {}
        tabla[mes_nombre][r.tipo_examen] = r.total

    return {
        "titulo": f"Mes x Tipo de Examen - {anio}",
        "anio": anio,
        "tabla": tabla
    }

@router.get("/codigo-atencion-periodo/{tipo_examen}")
def codigo_atencion_periodo(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    fecha: Optional[date] = None,
    mes: Optional[int] = Query(None, ge=1, le=12),
    anio: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        CodigoMAI.codigo,
        ExamenBase.atencion,
        func.count(ExamenBase.id).label("total")
    ).join(CodigoMAI).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    )

    periodo = ""
    if fecha:
        query = query.filter(ExamenBase.fecha_realizacion == fecha)
        periodo = fecha.strftime("%d/%m/%Y")
    elif mes and anio:
        query = query.filter(ExamenBase.mes_realizacion == mes, ExamenBase.anio_realizacion == anio)
        periodo = f"{mes}/{anio}"
    elif anio:
        query = query.filter(ExamenBase.anio_realizacion == anio)
        periodo = str(anio)

    query = query.group_by(CodigoMAI.codigo, ExamenBase.atencion)
    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.codigo not in tabla:
            tabla[r.codigo] = {}
        tabla[r.codigo][r.atencion] = r.total

    return {
        "titulo": f"Código x Atención - {tipo_examen} - {periodo}",
        "periodo": periodo,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/procedencia-atencion-periodo/{tipo_examen}")
def procedencia_atencion_periodo(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    fecha: Optional[date] = None,
    mes: Optional[int] = Query(None, ge=1, le=12),
    anio: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        Procedencia.nombre,
        ExamenBase.atencion,
        func.count(ExamenBase.id).label("total")
    ).join(Procedencia).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    )

    periodo = ""
    if fecha:
        query = query.filter(ExamenBase.fecha_realizacion == fecha)
        periodo = fecha.strftime("%d/%m/%Y")
    elif mes and anio:
        query = query.filter(ExamenBase.mes_realizacion == mes, ExamenBase.anio_realizacion == anio)
        periodo = f"{mes}/{anio}"
    elif anio:
        query = query.filter(ExamenBase.anio_realizacion == anio)
        periodo = str(anio)

    query = query.group_by(Procedencia.nombre, ExamenBase.atencion)
    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.nombre not in tabla:
            tabla[r.nombre] = {}
        tabla[r.nombre][r.atencion] = r.total

    return {
        "titulo": f"Procedencia x Atención - {tipo_examen} - {periodo}",
        "periodo": periodo,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/codigo-prevision-periodo/{tipo_examen}")
def codigo_prevision_periodo(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    fecha: Optional[date] = None,
    mes: Optional[int] = Query(None, ge=1, le=12),
    anio: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        CodigoMAI.codigo,
        Prevision.nombre,
        func.count(ExamenBase.id).label("total")
    ).join(CodigoMAI).join(Prevision).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    )

    periodo = ""
    if fecha:
        query = query.filter(ExamenBase.fecha_realizacion == fecha)
        periodo = fecha.strftime("%d/%m/%Y")
    elif mes and anio:
        query = query.filter(ExamenBase.mes_realizacion == mes, ExamenBase.anio_realizacion == anio)
        periodo = f"{mes}/{anio}"
    elif anio:
        query = query.filter(ExamenBase.anio_realizacion == anio)
        periodo = str(anio)

    query = query.group_by(CodigoMAI.codigo, Prevision.nombre)
    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.codigo not in tabla:
            tabla[r.codigo] = {}
        tabla[r.codigo][r.nombre] = r.total

    return {
        "titulo": f"Código x Previsión - {tipo_examen} - {periodo}",
        "periodo": periodo,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/atencion-contrato-periodo/{tipo_examen}")
def atencion_contrato_periodo(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    fecha: Optional[date] = None,
    mes: Optional[int] = Query(None, ge=1, le=12),
    anio: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        ExamenBase.atencion,
        ExamenBase.contrato,
        func.count(ExamenBase.id).label("total")
    ).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.deleted_at.is_(None)
    )

    periodo = ""
    if fecha:
        query = query.filter(ExamenBase.fecha_realizacion == fecha)
        periodo = fecha.strftime("%d/%m/%Y")
    elif mes and anio:
        query = query.filter(ExamenBase.mes_realizacion == mes, ExamenBase.anio_realizacion == anio)
        periodo = f"{mes}/{anio}"
    elif anio:
        query = query.filter(ExamenBase.anio_realizacion == anio)
        periodo = str(anio)

    query = query.group_by(ExamenBase.atencion, ExamenBase.contrato)
    resultados = query.all()

    tabla = {}
    for r in resultados:
        if r.atencion not in tabla:
            tabla[r.atencion] = {}
        tabla[r.atencion][r.contrato] = r.total

    return {
        "titulo": f"Atención x Contrato - {tipo_examen} - {periodo}",
        "periodo": periodo,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/mes-atencion-anio/{tipo_examen}")
def mes_atencion_anio(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        ExamenBase.mes_realizacion,
        ExamenBase.atencion,
        func.count(ExamenBase.id).label("total")
    ).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.mes_realizacion, ExamenBase.atencion).order_by(ExamenBase.mes_realizacion)

    resultados = query.all()
    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if mes_nombre not in tabla:
            tabla[mes_nombre] = {}
        tabla[mes_nombre][r.atencion] = r.total

    return {
        "titulo": f"Mes x Atención - {tipo_examen} - {anio}",
        "anio": anio,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/mes-mc-anio")
def mes_mc_anio_especifico(
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    """Gráfico opcional: Mes x Medio Contraste - Año Específico"""
    # Query: Mes x Si tiene MC / No tiene MC
    query = db.query(
        ExamenBase.mes_realizacion,
        ExamenTAC.medio_contraste,
        func.count(ExamenBase.id).label('total')
    ).join(ExamenTAC).filter(
        ExamenBase.tipo_examen == "TAC",
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(
        ExamenBase.mes_realizacion,
        ExamenTAC.medio_contraste
    ).order_by(ExamenBase.mes_realizacion)
    
    resultados = query.all()
    
    # Estructura: {mes_nombre: {"Con MC": X, "Sin MC": Y}}
    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if mes_nombre not in tabla:
            tabla[mes_nombre] = {}
        mc_label = "Con MC" if r.medio_contraste else "Sin MC"
        tabla[mes_nombre][mc_label] = r.total
    
    return {
        "titulo": f"Mes x Medio Contraste - {anio}",
        "anio": anio,
        "tabla": tabla
    }

@router.get("/mes-atencion-contrato/{tipo_examen}")
def mes_atencion_contrato(
    tipo_examen: str = Path(..., pattern="^(TAC|RX|ECO)$"),
    anio: int = Query(...),
    contrato: str = Query(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        ExamenBase.mes_realizacion,
        ExamenBase.atencion,
        func.count(ExamenBase.id).label("total")
    ).filter(
        ExamenBase.tipo_examen == tipo_examen,
        ExamenBase.anio_realizacion == anio,
        ExamenBase.contrato == contrato,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.mes_realizacion, ExamenBase.atencion).order_by(ExamenBase.mes_realizacion)

    resultados = query.all()

    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if mes_nombre not in tabla:
            tabla[mes_nombre] = {}
        tabla[mes_nombre][r.atencion] = r.total

    return {
        "titulo": f"Mes x Atención - {tipo_examen} - {anio} -{contrato}",
        "anio": anio,
        "contrato": contrato,
        "tipo_examen": tipo_examen,
        "tabla": tabla
    }

@router.get("/mes-tipo-examen-anio")
def mes_tipo_examen_anio_especifico(
    anio: int = Query(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    query = db.query(
        ExamenBase.mes_realizacion,
        ExamenBase.tipo_examen,
        func.count(ExamenBase.id).label("total")
    ).filter(
        ExamenBase.anio_realizacion == anio,
        ExamenBase.deleted_at.is_(None)
    ).group_by(ExamenBase.mes_realizacion, ExamenBase.tipo_examen).order_by(ExamenBase.mes_realizacion)

    resultados = query.all()

    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if mes_nombre not in tabla:
            tabla[mes_nombre] = {}
        tabla[mes_nombre][r.tipo_examen] = r.total

    return {
        "titulo": f"Mes x Tipo de Examen - {anio}",
        "anio": anio,
        "tabla": tabla
    }

@router.get("/mes-atencion-personal")
def mes_atencion_personal(
    tipo_examen: str = Query(..., pattern="^(TAC|RX|ECO)$"),
    anio: int = Query(...),
    contrato: str = Query(...),
    personal_id: int = Query(...),
    tipo_personal: str = Query(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user)
):
    from ..models.catalogos import PersonalMedico

    personal = db.query(PersonalMedico).filter(PersonalMedico.id == personal_id).first()
    if not personal:
        raise HTTPException(status_code=404, detail="Personal no encontrado")
    
    if tipo_examen == "TAC":
        if tipo_personal == "MEDICO":
            query = db.query(
                ExamenBase.mes_realizacion,
                ExamenBase.atencion,
                func.count(ExamenBase.id).label("total")
            ).join(ExamenTAC).filter(
                ExamenTAC.medico_solicitante_id == personal_id,
                ExamenBase.anio_realizacion == anio,
                ExamenBase.contrato == contrato,
                ExamenBase.deleted_at.is_(None)
            )
        elif tipo_personal == "TM":
            query = db.query(
                ExamenBase.mes_realizacion,
                ExamenBase.atencion,
                func.count(ExamenBase.id).label("total")
            ).join(ExamenTAC).filter(
                ExamenTAC.tm_id == personal_id,
                ExamenBase.anio_realizacion == anio,
                ExamenBase.contrato == contrato,
                ExamenBase.deleted_at.is_(None)
            )
        elif tipo_personal == "TP":
            query = db.query(
                ExamenBase.mes_realizacion,
                ExamenBase.atencion,
                func.count(ExamenBase.id).label("total")
            ).join(ExamenTAC).filter(
                ExamenTAC.tp_id == personal_id,
                ExamenBase.anio_realizacion == anio,
                ExamenBase.contrato == contrato,
                ExamenBase.deleted_at.is_(None)
            )
        elif tipo_personal == "SECRETARIA":
            query = db.query(
                ExamenBase.mes_realizacion,
                ExamenBase.atencion,
                func.count(ExamenBase.id).label("total")
            ).join(ExamenTAC).filter(
                ExamenTAC.secretaria_id == personal_id,
                ExamenBase.anio_realizacion == anio,
                ExamenBase.contrato == contrato,
                ExamenBase.deleted_at.is_(None)
            )
    elif tipo_examen == "RX":
        from ..models.examen_rx import ExamenRX
        query = db.query(
            ExamenBase.mes_realizacion,
            ExamenBase.atencion,
            func.count(ExamenBase.id).label("total")
        ).join(ExamenRX).filter(
            ExamenRX.tm_tp_id == personal_id,
            ExamenBase.anio_realizacion == anio,
            ExamenBase.contrato == contrato,
            ExamenBase.deleted_at.is_(None)
        ) 
    elif tipo_examen == "ECO":
        from ..models.examen_eco import ExamenECO
        if tipo_personal == "REALIZADO":
            query = db.query(
                ExamenBase.mes_realizacion,
                ExamenBase.atencion,
                func.count(ExamenBase.id).label("total")
            ).join(ExamenECO).filter(
                ExamenECO.realizado_id == personal_id,
                ExamenBase.anio_realizacion == anio,
                ExamenBase.contrato == contrato,
                ExamenBase.deleted_at.is_(None)
            )
        elif tipo_personal == "TRANSCRIBE":
            query = db.query(
                ExamenBase.mes_realizacion,
                ExamenBase.atencion,
                func.count(ExamenBase.id).label("total")
            ).join(ExamenECO).filter(
                ExamenECO.transcribe_id == personal_id,
                ExamenBase.anio_realizacion == anio,
                ExamenBase.contrato == contrato,
                ExamenBase.deleted_at.is_(None)
            )
    
    query = query.group_by(ExamenBase.mes_realizacion, ExamenBase.atencion).order_by(ExamenBase.mes_realizacion)
    resultados = query.all()

    tabla = {}
    for r in resultados:
        mes_nombre = MESES[r.mes_realizacion - 1] if 1 <= r.mes_realizacion <= 12 else f"Mes {r.mes_realizacion}"
        if mes_nombre not in tabla:
            tabla[mes_nombre] = {}
        tabla[mes_nombre][r.atencion] = r.total

    return {
        "titulo": f"Mes x Atención - {tipo_examen} - {anio} - {contrato} - {personal.nombre} ({tipo_personal})",
        "anio": anio,
        "contrato": contrato,
        "tipo_examen": tipo_examen,
        "personal": personal.nombre,
        "tabla": tabla
    }


@router.get("/exportar-graficos-excel")
def exportar_graficos_excel(
    current_user: Usuario = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Exporta todos los gráficos obligatorios a Excel con tablas pivote
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Lista de gráficos a exportar - CORREGIDO
    graficos = [
        ("Código-Atención TAC", codigo_atencion_mes_actual("TAC", db)),
        ("Código-Atención RX", codigo_atencion_mes_actual("RX", db)),
        ("Código-Atención ECO", codigo_atencion_mes_actual("ECO", db)),
        ("Procedencia-Atención TAC", procedencia_atencion_mes_actual("TAC", db)),
        ("Procedencia-Atención RX", procedencia_atencion_mes_actual("RX", db)),
        ("Procedencia-Atención ECO", procedencia_atencion_mes_actual("ECO", db)),
        ("Código-Previsión TAC", codigo_prevision_mes_actual("TAC", db)),
        ("Código-Previsión RX", codigo_prevision_mes_actual("RX", db)),
        ("Código-Previsión ECO", codigo_prevision_mes_actual("ECO", db)),
        ("Atención-Contrato TAC", atencion_contrato_mes_actual("TAC", db)),
        ("Atención-Contrato RX", atencion_contrato_mes_actual("RX", db)),
        ("Atención-Contrato ECO", atencion_contrato_mes_actual("ECO", db)),
        ("Mes-Año TAC", mes_anio_tipo_examen("TAC", db)),
        ("Mes-Año RX", mes_anio_tipo_examen("RX", db)),
        ("Mes-Año ECO", mes_anio_tipo_examen("ECO", db)),
        ("Mes-MC Año Actual", mes_anio_actual(db)),  # CORREGIDO: era mes_mc_anio_actual
        ("Mes-Año-MC", mes_anio_mc(db)),
        ("Mes-Tipo Examen", mes_tipo_examen_anio_actual(db))
    ]
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    total_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for nombre_hoja, grafico_data in graficos:
        nombre_corto = nombre_hoja[:31]
        ws = wb.create_sheet(title=nombre_corto)
        
        tabla = grafico_data.get("tabla", {})
        
        if not tabla or not isinstance(tabla, dict):
            ws.append(["Sin datos"])
            continue
        
        # Obtener filas (excluyendo "Total" si existe)
        filas = [f for f in tabla.keys() if f != "Total"]
        if not filas:
            ws.append(["Sin datos"])
            continue
        
        # Obtener columnas verificando que cada fila sea un diccionario
        columnas = []
        for fila in filas:
            if isinstance(tabla[fila], dict):
                columnas.extend([c for c in tabla[fila].keys() if c != "Total"])
        
        # Eliminar duplicados manteniendo orden
        columnas = list(dict.fromkeys(columnas))
        
        if not columnas:
            ws.append(["Sin datos"])
            continue
        
        # Escribir encabezado
        encabezados = [""] + columnas + ["TOTAL"]
        ws.append(encabezados)
        
        # Aplicar estilo al encabezado
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Escribir datos
        for fila in filas:
            # Ya viene como nombre de mes desde el backend, no convertir
            nombre_fila = str(fila)
            
            valores = [nombre_fila]
            total_fila = 0
            
            if isinstance(tabla[fila], dict):
                for col in columnas:
                    valor = tabla[fila].get(col, 0)
                    valores.append(valor)
                    total_fila += valor
            else:
                # Si tabla[fila] no es dict, rellenar con ceros
                valores.extend([0] * len(columnas))
            
            valores.append(total_fila)
            ws.append(valores)
            
            for cell in ws[ws.max_row]:
                cell.border = border
                if cell.column == 1:
                    cell.font = Font(bold=True)
                if cell.column == len(encabezados):
                    cell.font = Font(bold=True)
        
        # Fila TOTAL
        fila_total = ["TOTAL"]
        gran_total = 0
        
        for col in columnas:
            total_col = 0
            for f in filas:
                if isinstance(tabla[f], dict):
                    total_col += tabla[f].get(col, 0)
            fila_total.append(total_col)
            gran_total += total_col
        
        fila_total.append(gran_total)
        ws.append(fila_total)
        
        for cell in ws[ws.max_row]:
            cell.fill = total_fill
            cell.font = total_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Ajustar ancho de columnas
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=graficos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        }
    )